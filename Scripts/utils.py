import os
import shutil
import openpyxl
import xlrd

TEMP_DIR = "temp_files"

def split_message(text, max_length=4096):
    """
    Разбивает текст на части, не превышающие max_length символов
    :param text: Исходный текст
    :param max_length: Максимальная длина части
    :return: Список частей
    """
    lines = text.split("\n")
    chunks = []
    current_chunk = ""

    for line in lines:
        if len(current_chunk) + len(line) + 1 <= max_length:
            current_chunk += line + "\n"
        else:
            chunks.append(current_chunk.strip())
            current_chunk = line + "\n"
    if current_chunk.strip():
        chunks.append(current_chunk.strip())

    return chunks

def download_and_convert_xls(bot, file_id: str, temp_dir: str, file_name: str) -> str:
    """
    Загружает .xls файл, конвертирует его в .xlsx и возвращает путь к .xlsx файлу

    :param bot: объект бота Telegram
    :param file_id: Идентификатор файла из Telegram
    :param temp_dir: Директория для временных файлов
    :param file_name: Имя файла
    :return: Путь к .xlsx файлу
    """
    # Download .xls file
    file_info = bot.get_file(file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    xls_file_path = os.path.join(temp_dir, file_name)

    with open(xls_file_path, "wb") as new_file:
        new_file.write(downloaded_file)

    # Convert path to .xlsx
    xlsx_file_name = file_name.replace(".xls", ".xlsx")
    xlsx_file_path = os.path.join(temp_dir, xlsx_file_name)

    # Open .xls and convert to .xlsx
    try:
        wb_old = xlrd.open_workbook(xls_file_path, formatting_info=False)
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active

        for ws_old in wb_old.sheets():
            for row in range(ws_old.nrows):
                for col in range(ws_old.ncols):
                    ws_new.cell(row=row + 1, column=col + 1).value = ws_old.cell(row, col).value

        wb_new.save(xlsx_file_path)
    except Exception as e:
        raise RuntimeError(f"Ошибка при конвертации .xls в .xlsx: {e}")

    return xlsx_file_path

def clean_temp_folder():
    if os.path.exists(TEMP_DIR):
        shutil.rmtree(TEMP_DIR)
        os.makedirs(TEMP_DIR)